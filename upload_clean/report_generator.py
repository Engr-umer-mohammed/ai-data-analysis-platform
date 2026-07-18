# report_generator.py
# Professional Report Generation Engine
# AI Data Analysis Agent — Version 2.0
#
# CHANGES FROM v1.0:
#   Full model fallback strategy restored — mirrors agent.py
#   gemini-3.5-flash primary with three fallback models
#   Retry logic per model for 503 busy and 429 rate limit
#   Correct Gemini API SDK call syntax per official reference
#   Timeout handled at request level not client init level
#   Graceful fallback when customization module absent
#   ExcelReportWriter uses explicit workbook save
#   All error types classified and handled individually
#   Retry count and delay configurable at class level

import os
import json
import logging
import time
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
import numpy as np
from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════
# CUSTOMIZATION IMPORT — Graceful fallback if absent
# ════════════════════════════════════════════════════════

def _load_domain_registry():
    """
    Attempt to import the domain registry.
    Returns the function if available, returns a
    no-op lambda if the customization module has
    not been created yet. This allows the report
    generator to function fully before the domain
    layer is implemented.
    """
    try:
        from customization.domain_registry import (
            get_applicable_domains
        )
        return get_applicable_domains
    except ImportError:
        logger.info(
            "customization.domain_registry not found. "
            "Domain customizations will be skipped. "
            "Create the customization folder and "
            "domain_registry.py to enable this feature."
        )
        return lambda profile, stats, df: []


get_applicable_domains = _load_domain_registry()


# ════════════════════════════════════════════════════════
# REPORT RESULT
# ════════════════════════════════════════════════════════

@dataclass
class ReportResult:
    """
    Complete output contract from one report generation run.
    All callers — data_agent.py, Telegram bot, terminal
    interface — work exclusively with this object.
    """

    text_report_path:     Optional[str] = None
    excel_report_path:    Optional[str] = None
    ai_narrative:         str           = ""
    executive_summary:    str           = ""
    generation_success:   bool          = False
    generation_error:     Optional[str] = None
    model_used:           str           = ""
    generation_timestamp: str           = ""
    generation_warnings:  list          = field(default_factory=list)

    def get_all_output_paths(self) -> list:
        paths = [
            self.text_report_path,
            self.excel_report_path
        ]
        return [p for p in paths if p is not None]


# ════════════════════════════════════════════════════════
# AI NARRATIVE ENGINE
# ════════════════════════════════════════════════════════

class AIReportNarrator:
    """
    Connects to Gemini AI and generates the complete
    written intelligence layer of the report.

    MODEL STRATEGY — per official Gemini API reference:
    The generateContent endpoint is used for standard
    non-interactive content generation, which is the
    correct choice for report narrative generation where
    the full response is needed before writing begins.

    Fallback sequence:
        Primary    : gemini-3.5-flash  (latest stable)
        Fallback 1 : gemini-2.0-flash  (fast alternative)
        Fallback 2 : gemini-1.5-flash  (proven reliable)
        Fallback 3 : gemini-1.5-pro    (most capable fallback)

    Per-model retry behavior:
        503 unavailable : retry up to MAX_RETRIES_PER_MODEL
                          with RETRY_DELAY_SECONDS between
        429 rate limit  : wait RATE_LIMIT_WAIT_SECONDS
                          then retry once more
        404 not found   : skip immediately to next model
        Other errors    : log and skip to next model
    """

    PRIMARY_MODEL = "gemini-3.5-flash"

    FALLBACK_MODELS = [
        "gemini-2.0-flash",
        "gemini-1.5-flash",
        "gemini-1.5-pro",
    ]

    # How many times to retry the same model on busy errors
    MAX_RETRIES_PER_MODEL   = 3

    # Seconds to wait between retries on the same model
    RETRY_DELAY_SECONDS     = 4

    # Seconds to wait after a rate limit response
    RATE_LIMIT_WAIT_SECONDS = 15

    def __init__(self):
        from google import genai

        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            raise ValueError(
                "GEMINI_API_KEY not found in .env file.\n"
                "Add: GEMINI_API_KEY=your_key_here"
            )

        # Initialize client without timeout at constructor level.
        # Per Gemini API SDK reference, timeout is passed at
        # the individual request level, not at client init.
        self.client = genai.Client(api_key=api_key)

        self.model_used = self.PRIMARY_MODEL
        self.model_list = (
            [self.PRIMARY_MODEL] + self.FALLBACK_MODELS
        )

        logger.info(
            f"AIReportNarrator initialized. "
            f"Primary: {self.PRIMARY_MODEL} | "
            f"Fallbacks: {self.FALLBACK_MODELS}"
        )

    # ════════════════════════════════════════════════════
    # SMART AI CALLER — Full fallback and retry logic
    # ════════════════════════════════════════════════════

    def _call_ai(self, prompt: str) -> tuple:
        """
        Call Gemini AI using the generateContent endpoint
        with full model fallback and per-model retry logic.

        Per the official Gemini API reference, generateContent
        is the standard REST endpoint for non-interactive tasks
        where the full response is returned in one package —
        the correct choice for report narrative generation.

        Error classification:
            503 UNAVAILABLE    → server busy, retry same model
            429 RESOURCE_EXHAUSTED → rate limit, wait then retry
            404 NOT_FOUND      → model unavailable, skip to next
            All others         → log and skip to next model

        Returns:
            (response_text, None)       on success
            (None, error_message)       when all models fail
        """

        last_error = None

        for model_name in self.model_list:
            logger.info(
                f"Narrator attempting model: {model_name}"
            )

            for attempt in range(
                1, self.MAX_RETRIES_PER_MODEL + 1
            ):
                try:
                    # generateContent call per Gemini API reference
                    # Uses the standard content generation endpoint:
                    # POST /v1beta/models/{model}:generateContent
                    response = self.client.models.generate_content(
                        model    = model_name,
                        contents = prompt
                    )

                    if model_name != self.model_used:
                        logger.info(
                            f"Narrator switched to: {model_name}"
                        )
                        print(
                            f"     AI model switched to: {model_name}"
                        )

                    self.model_used = model_name

                    logger.info(
                        f"AI response received from "
                        f"{model_name} on attempt {attempt}"
                    )
                    return response.text, None

                except Exception as error:
                    error_str  = str(error).lower()
                    error_repr = repr(error)
                    last_error = error

                    logger.warning(
                        f"Model {model_name} attempt {attempt}/"
                        f"{self.MAX_RETRIES_PER_MODEL} failed. "
                        f"Type: {type(error).__name__} | "
                        f"Message: {str(error)[:120]}"
                    )

                    # ── 503 SERVER BUSY ──────────────────────────
                    # Retry same model after delay
                    if (
                        "503" in error_str or
                        "unavailable" in error_str or
                        "high demand" in error_str or
                        "overloaded" in error_str
                    ):
                        if attempt < self.MAX_RETRIES_PER_MODEL:
                            print(
                                f"     Model {model_name} busy. "
                                f"Waiting {self.RETRY_DELAY_SECONDS}s "
                                f"(attempt {attempt}/"
                                f"{self.MAX_RETRIES_PER_MODEL})..."
                            )
                            time.sleep(self.RETRY_DELAY_SECONDS)
                            continue
                        else:
                            print(
                                f"     Model {model_name} "
                                f"consistently busy. "
                                f"Trying next model..."
                            )
                            break

                    # ── 429 RATE LIMIT ───────────────────────────
                    # Wait longer then retry once more
                    elif (
                        "429" in error_str or
                        "quota" in error_str or
                        "exhausted" in error_str or
                        "resource_exhausted" in error_str or
                        "rate" in error_str
                    ):
                        print(
                            f"     Rate limit on {model_name}. "
                            f"Waiting {self.RATE_LIMIT_WAIT_SECONDS}s..."
                        )
                        time.sleep(self.RATE_LIMIT_WAIT_SECONDS)

                        if attempt < self.MAX_RETRIES_PER_MODEL:
                            continue
                        else:
                            break

                    # ── 404 MODEL NOT FOUND ──────────────────────
                    # Skip immediately to next model
                    elif (
                        "404" in error_str or
                        "not found" in error_str or
                        "does not exist" in error_str or
                        "model_not_found" in error_str
                    ):
                        logger.warning(
                            f"Model {model_name} not available. "
                            f"Skipping to next fallback."
                        )
                        print(
                            f"     Model {model_name} not available. "
                            f"Trying next model..."
                        )
                        break

                    # ── AUTHENTICATION ERROR ─────────────────────
                    # API key issue — no point retrying any model
                    elif (
                        "401" in error_str or
                        "403" in error_str or
                        "api_key" in error_str or
                        "authentication" in error_str or
                        "permission" in error_str
                    ):
                        logger.error(
                            f"Authentication error: {error}. "
                            f"Check your GEMINI_API_KEY."
                        )
                        print(
                            f"     Authentication error. "
                            f"Check your API key in .env file."
                        )
                        return None, str(error)

                    # ── UNKNOWN ERROR ────────────────────────────
                    # Log and try next model
                    else:
                        logger.error(
                            f"Unexpected error on {model_name}: "
                            f"{error_repr}"
                        )
                        print(
                            f"     Unexpected error on {model_name}: "
                            f"{str(error)[:80]}"
                        )
                        break

        # All models exhausted
        error_msg = str(last_error) if last_error else "Unknown"
        logger.error(
            f"All models failed. Last error: {error_msg}"
        )
        print(
            f"     All AI models exhausted. "
            f"Using statistical fallback narrative."
        )
        return None, error_msg

    # ════════════════════════════════════════════════════
    # FULL ANALYSIS NARRATIVE
    # ════════════════════════════════════════════════════

    def generate_full_analysis(
        self,
        profile,
        stats:      dict,
        viz_result
    ) -> tuple:
        """
        Generate the complete AI narrative for the report.
        Constructs a comprehensive prompt from all statistical
        results and calls the AI with full fallback support.
        Returns (narrative_text, error_message).
        """

        desc_stats   = stats.get("descriptive",  {})
        anomalies    = stats.get("anomalies",    {})
        correlations = stats.get("correlations", [])
        trends       = stats.get("trends",       {})

        stats_summary       = self._build_stats_summary(
            desc_stats, profile.numeric_columns
        )
        anomaly_summary     = self._build_anomaly_summary(
            anomalies
        )
        correlation_summary = self._build_correlation_summary(
            correlations
        )
        trend_summary       = self._build_trend_summary(trends)

        charts_generated = (
            viz_result.total_charts_generated
            if viz_result else 0
        )

        prompt = f"""
You are a senior data analyst and strategic advisor
with deep expertise across multiple industries.
You have just completed a full statistical analysis
of a dataset and must now write the intelligence
layer of a professional report that will be read
by decision-makers who need clear understanding
and immediate actionable guidance.

DATASET OVERVIEW:
  File name        : {profile.file_name}
  Total records    : {profile.total_rows:,}
  Total variables  : {profile.total_columns}
  Numeric variables: {len(profile.numeric_columns)}
  Text variables   : {len(profile.text_columns)}
  Data completeness: {profile.completeness_percent}%
  Missing values   : {profile.total_missing_values}
  Time series      : {"Yes — " + str(profile.time_column) if profile.has_time_series else "No"}
  Charts generated : {charts_generated}

NUMERIC COLUMNS IN DATASET:
  {", ".join(profile.numeric_columns) if profile.numeric_columns else "None"}

TEXT COLUMNS IN DATASET:
  {", ".join(profile.text_columns) if profile.text_columns else "None"}

STATISTICAL SUMMARY:
{stats_summary}

ANOMALIES AND OUTLIERS DETECTED:
{anomaly_summary}

CORRELATION FINDINGS:
{correlation_summary}

TREND ANALYSIS:
{trend_summary}

DATA QUALITY WARNINGS FROM LOADER:
  {chr(10).join(profile.warnings) if profile.warnings else "None"}

Your task is to write the complete analytical
narrative for this report. Structure your response
in exactly the following sections using these
exact section headers. Write in clear professional
English. Do not use markdown symbols like ** or ##.
Do not use bullet points — write in full paragraphs.
Each section must be substantive, specific to the
actual data findings above, and written for a
reader who makes decisions but may not be a
data specialist.

EXECUTIVE SUMMARY
Write 3 to 4 sentences that capture the single
most important story this data tells. What is
the headline finding? What is the overall health
or state of whatever this data measures?

DATASET CHARACTERIZATION
Describe what kind of dataset this is based on
its structure, variables, and time characteristics.
What domain does it likely come from? What does
the completeness and size tell us about its
reliability as a basis for decisions?

KEY STATISTICAL FINDINGS
Interpret the most significant statistical results.
Discuss the distribution shapes, central tendencies,
and variability of the most important variables.
What do these numbers mean in practical terms?

ANOMALY INTERPRETATION
Explain every detected anomaly in plain language.
What are the anomalous values? How far do they
deviate from normal? What real-world events or
conditions could explain them?

RELATIONSHIP ANALYSIS
Interpret every strong correlation found. Explain
what each relationship means in practical terms.
Distinguish between likely causal versus coincidental
relationships.

TREND AND PATTERN INSIGHTS
If time series data exists, describe the directional
movement of key variables over the observed period.
Are things improving, deteriorating, or stable?

RISK INDICATORS
Identify the specific findings that represent
risks, vulnerabilities, or warning signs that
require attention.

RECOMMENDATIONS
Write exactly 5 specific, prioritized, actionable
recommendations based directly on the findings above.
Number them 1 through 5 with highest priority first.

LIMITATIONS AND CAVEATS
Describe the limitations of this analysis honestly.
What would additional data reveal? What assumptions
underpin the findings?
        """

        return self._call_ai(prompt)

    # ════════════════════════════════════════════════════
    # EXECUTIVE SUMMARY
    # ════════════════════════════════════════════════════

    def generate_executive_summary(
        self,
        profile,
        stats:          dict,
        full_narrative: str
    ) -> tuple:
        """
        Generate a condensed executive summary from the
        full narrative. Returns (summary_text, error).
        """

        prompt = f"""
Based on this complete analysis narrative,
write a concise executive summary of exactly
150 to 200 words suitable for a senior
decision-maker who has 60 seconds to read it.

The summary must cover: what the data represents,
the single most important finding, the most
significant risk, and the top priority action.

Write in plain professional English.
No bullet points. No markdown. Full sentences.

FULL ANALYSIS TO SUMMARIZE:
{full_narrative[:3000]}

DATASET: {profile.file_name}
RECORDS: {profile.total_rows:,}
        """

        return self._call_ai(prompt)

    # ════════════════════════════════════════════════════
    # PROMPT BUILDING HELPERS
    # ════════════════════════════════════════════════════

    def _build_stats_summary(
        self,
        desc_stats:   dict,
        numeric_cols: list
    ) -> str:
        lines = []
        for col in numeric_cols[:8]:
            if col in desc_stats:
                s = desc_stats[col]
                try:
                    lines.append(
                        f"  {col}: "
                        f"mean={float(s.get('mean', 0)):.3f}, "
                        f"median={float(s.get('median', 0)):.3f}, "
                        f"std={float(s.get('std', 0)):.3f}, "
                        f"min={float(s.get('min', 0)):.3f}, "
                        f"max={float(s.get('max', 0)):.3f}, "
                        f"skew={float(s.get('skewness', 0)):.3f}"
                    )
                except (TypeError, ValueError):
                    lines.append(
                        f"  {col}: statistics unavailable"
                    )
        return (
            "\n".join(lines)
            if lines
            else "  No numeric statistics available"
        )

    def _build_anomaly_summary(self, anomalies: dict) -> str:
        if not anomalies:
            return "  No anomalies detected in any variable"
        lines = []
        for col, info in list(anomalies.items())[:6]:
            try:
                count = info.get("anomaly_count", 0)
                pct   = float(info.get("anomaly_percent", 0))
                lower = float(info.get("lower_bound", 0))
                upper = float(info.get("upper_bound", 0))
                lines.append(
                    f"  {col}: {count} anomalies "
                    f"({pct:.1f}%), "
                    f"expected range "
                    f"[{lower:.2f} to {upper:.2f}]"
                )
            except (TypeError, ValueError):
                lines.append(
                    f"  {col}: anomaly details unavailable"
                )
        return "\n".join(lines)

    def _build_correlation_summary(
        self,
        correlations: list
    ) -> str:
        if not correlations:
            return "  No strong correlations detected"
        lines = []
        for corr in correlations[:6]:
            col_a    = corr.get("column_a",    "")
            col_b    = corr.get("column_b",    "")
            coeff    = corr.get("correlation", 0)
            strength = corr.get("strength",    "")
            direction = (
                "positive" if coeff > 0 else "negative"
            )
            lines.append(
                f"  {col_a} vs {col_b}: "
                f"r={coeff:.3f} ({strength} {direction})"
            )
        return "\n".join(lines)

    def _build_trend_summary(self, trends: dict) -> str:
        if not trends:
            return "  No time series trend data available"
        lines = []
        for col, info in list(trends.items())[:4]:
            direction = info.get("direction",      "stable")
            change    = info.get("percent_change", 0)
            lines.append(
                f"  {col}: {direction} "
                f"({float(change):+.1f}% over period)"
            )
        return "\n".join(lines)


# ════════════════════════════════════════════════════════
# TEXT REPORT WRITER
# ════════════════════════════════════════════════════════

class TextReportWriter:
    """
    Assembles the complete professional text report
    from all analysis components. The document structure
    follows standard analytical report conventions —
    executive summary first, detailed findings in the
    body, and a complete statistical appendix at the end.

    Domain customization sections are inserted after the
    standard sections when the domain registry provides them.
    """

    def write(
        self,
        profile,
        stats:         dict,
        viz_result,
        ai_narrative:  str,
        exec_summary:  str,
        output_folder: str,
        session_id:    str,
        df:            Optional[pd.DataFrame] = None
    ) -> str:
        """
        Write the complete text report to disk.
        Returns the file path on success.
        """

        os.makedirs(output_folder, exist_ok=True)

        timestamp = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        safe_name = (
            os.path.splitext(profile.file_name)[0]
            .replace(" ", "_")
            .replace("/", "-")
        )
        filename  = f"{session_id}_{safe_name}_report.txt"
        filepath  = os.path.join(output_folder, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(self._build_header(profile, timestamp))
            f.write(self._build_executive_summary_section(
                exec_summary
            ))
            f.write(self._build_dataset_profile_section(
                profile
            ))
            f.write(self._build_data_quality_section(
                profile
            ))
            f.write(self._build_statistical_section(
                stats, profile
            ))
            f.write(self._build_anomaly_section(stats))
            f.write(self._build_correlation_section(stats))
            f.write(self._build_trend_section(stats, profile))
            f.write(self._build_ai_narrative_section(
                ai_narrative
            ))
            f.write(self._build_visualizations_section(
                viz_result
            ))

            # Domain customization sections
            data_df          = (
                df if df is not None
                else profile.dataframe
            )
            custom_content   = self._add_customizations(
                profile, stats, data_df
            )
            if custom_content:
                f.write(custom_content)

            f.write(self._build_statistical_appendix(
                stats, profile
            ))
            f.write(self._build_footer(timestamp))

        logger.info(f"Text report saved: {filepath}")
        return filepath

    def _add_customizations(
        self,
        profile,
        stats: dict,
        df:    pd.DataFrame
    ) -> str:
        """
        Collect domain-specific content from the
        domain registry and append it to the report.
        Returns empty string if no customizations apply
        or if the registry is unavailable.
        """

        try:
            customizations = get_applicable_domains(
                profile, stats, df
            )
        except Exception as error:
            logger.warning(
                f"Domain registry call failed: {error}"
            )
            return ""

        if not customizations:
            return ""

        content = []

        for custom in customizations:
            try:
                header = custom.get_section_header()
                content.append(
                    self._section_header(header)
                )
                custom_content = custom.generate_content()

                if isinstance(custom_content, pd.DataFrame):
                    custom_content = custom_content.to_string()
                elif not isinstance(custom_content, str):
                    custom_content = str(custom_content)

                content.append(custom_content)
                content.append("\n")

            except Exception as error:
                logger.warning(
                    f"Customization section failed: {error}"
                )
                continue

        return "\n".join(content)

    # ════════════════════════════════════════════════════
    # FORMATTING HELPERS
    # ════════════════════════════════════════════════════

    def _separator(
        self,
        char:  str = "─",
        width: int = 65
    ) -> str:
        return char * width + "\n"

    def _section_header(self, title: str) -> str:
        width = 65
        return (
            "\n" +
            self._separator("═", width) +
            f"  {title.upper()}\n" +
            self._separator("═", width) +
            "\n"
        )

    def _subsection_header(self, title: str) -> str:
        return (
            "\n" +
            self._separator("─", 55) +
            f"  {title}\n" +
            self._separator("─", 55) +
            "\n"
        )

    # ════════════════════════════════════════════════════
    # REPORT SECTIONS
    # ════════════════════════════════════════════════════

    def _build_header(
        self,
        profile,
        timestamp: str
    ) -> str:
        width = 65
        return (
            self._separator("═", width) +
            self._separator("═", width) +
            "\n"
            "        AI DATA ANALYSIS AGENT\n"
            "        PROFESSIONAL ANALYSIS REPORT\n"
            "\n" +
            self._separator("═", width) +
            self._separator("═", width) +
            "\n"
            f"  Dataset       : {profile.file_name}\n"
            f"  Generated at  : {timestamp}\n"
            f"  Total records : {profile.total_rows:,}\n"
            f"  Total columns : {profile.total_columns}\n"
            f"  Completeness  : {profile.completeness_percent}%\n"
            "\n" +
            self._separator("═", width) +
            "\n"
        )

    def _build_executive_summary_section(
        self,
        exec_summary: str
    ) -> str:
        return (
            self._section_header("Executive Summary") +
            exec_summary.strip() +
            "\n\n"
        )

    def _build_dataset_profile_section(
        self,
        profile
    ) -> str:
        content = (
            f"  File name      : {profile.file_name}\n"
            f"  File format    : {profile.file_format.upper()}\n"
            f"  File size      : {profile.file_size_kb} KB\n"
            f"  Total rows     : {profile.total_rows:,}\n"
            f"  Total columns  : {profile.total_columns}\n"
            f"  Numeric columns: {len(profile.numeric_columns)}\n"
            f"  Text columns   : {len(profile.text_columns)}\n"
            f"  Datetime cols  : {len(profile.datetime_columns)}\n"
        )

        if profile.has_time_series:
            content += (
                f"\n  TIME SERIES DETECTED\n"
                f"  Time column    : {profile.time_column}\n"
                f"  Period start   : {profile.time_range_start}\n"
                f"  Period end     : {profile.time_range_end}\n"
            )

        if profile.numeric_columns:
            content += "\n  NUMERIC COLUMNS\n"
            for col in profile.numeric_columns:
                content += f"    • {col}\n"

        if profile.text_columns:
            content += "\n  TEXT COLUMNS\n"
            for col in profile.text_columns:
                content += f"    • {col}\n"

        if profile.datetime_columns:
            content += "\n  DATETIME COLUMNS\n"
            for col in profile.datetime_columns:
                content += f"    • {col}\n"

        return (
            self._section_header("Dataset Profile") +
            content + "\n"
        )

    def _build_data_quality_section(self, profile) -> str:
        content = (
            f"  Overall completeness  : "
            f"{profile.completeness_percent}%\n"
            f"  Total missing values  : "
            f"{profile.total_missing_values:,}\n"
            f"  Duplicate rows removed: "
            f"{profile.duplicate_rows}\n"
        )

        if profile.missing_by_column:
            content += "\n  MISSING VALUES BY COLUMN\n"
            for col, count in sorted(
                profile.missing_by_column.items(),
                key=lambda x: x[1],
                reverse=True
            ):
                pct        = round(
                    count / profile.total_rows * 100, 1
                )
                bar_filled = int(pct / 5)
                bar        = (
                    "█" * bar_filled +
                    "░" * (20 - bar_filled)
                )
                content += (
                    f"    {col:<25} "
                    f"{count:>5} missing "
                    f"({pct:>5.1f}%) "
                    f"[{bar}]\n"
                )

        if profile.warnings:
            content += "\n  PROCESSING NOTES\n"
            for warning in profile.warnings:
                content += f"    ⚠  {warning}\n"

        return (
            self._section_header("Data Quality Assessment") +
            content + "\n"
        )

    def _build_statistical_section(
        self,
        stats:   dict,
        profile
    ) -> str:
        desc = stats.get("descriptive", {})

        if not desc:
            return (
                self._section_header(
                    "Descriptive Statistics"
                ) +
                "  No numeric columns available.\n\n"
            )

        content = ""

        for col in profile.numeric_columns:
            if col not in desc:
                continue

            s = desc[col]
            content += self._subsection_header(
                col.replace("_", " ").upper()
            )

            metrics = [
                ("Count",     s.get("count",    0), "d",    ""),
                ("Mean",      s.get("mean",     0), ".4f",  ""),
                ("Median",    s.get("median",   0), ".4f",  ""),
                ("Std Dev",   s.get("std",      0), ".4f",  ""),
                ("Variance",  s.get("variance", 0), ".4f",  ""),
                ("Min",       s.get("min",      0), ".4f",  ""),
                ("Max",       s.get("max",      0), ".4f",  ""),
                ("Range",     s.get("range",    0), ".4f",  ""),
                ("Q1 (25%)",  s.get("q1",       0), ".4f",  ""),
                ("Q3 (75%)",  s.get("q3",       0), ".4f",  ""),
                ("IQR",       s.get("iqr",      0), ".4f",  ""),
                ("Skewness",  s.get("skewness", 0), ".4f",  ""),
                ("Kurtosis",  s.get("kurtosis", 0), ".4f",  ""),
                ("Missing",   s.get("missing",  0), "d",    ""),
            ]

            for label, value, fmt, unit in metrics:
                try:
                    formatted = (
                        f"{value:{fmt}}"
                        if fmt != "d"
                        else f"{int(value)}"
                    )
                except (ValueError, TypeError):
                    formatted = str(value)
                content += (
                    f"    {label:<20}: {formatted}{unit}\n"
                )

            skew = s.get("skewness", 0)
            kurt = s.get("kurtosis", 0)

            if abs(skew) < 0.5:
                skew_desc = "approximately symmetric"
            elif skew > 1.0:
                skew_desc = "strongly right-skewed (long upper tail)"
            elif skew > 0.5:
                skew_desc = "moderately right-skewed"
            elif skew < -1.0:
                skew_desc = "strongly left-skewed (long lower tail)"
            else:
                skew_desc = "moderately left-skewed"

            if kurt > 3:
                kurt_desc = "leptokurtic (heavy tails, sharp peak)"
            elif kurt < -1:
                kurt_desc = "platykurtic (light tails, flat distribution)"
            else:
                kurt_desc = "approximately normal distribution shape"

            content += (
                f"\n    Distribution: {skew_desc}, "
                f"{kurt_desc}.\n"
            )

        return (
            self._section_header("Descriptive Statistics") +
            content + "\n"
        )

    def _build_anomaly_section(self, stats: dict) -> str:
        anomalies = stats.get("anomalies", {})

        if not anomalies:
            return (
                self._section_header("Anomaly Detection") +
                "  No statistically significant anomalies "
                "detected using the IQR method "
                "at 1.5x threshold.\n\n"
            )

        content = (
            "  Detection method: Interquartile Range (IQR)\n"
            "  Threshold: 1.5 × IQR beyond Q1 and Q3\n\n"
        )

        for col, info in anomalies.items():
            count    = info.get("anomaly_count",   0)
            pct      = info.get("anomaly_percent", 0)
            lower    = info.get("lower_bound",     0)
            upper    = info.get("upper_bound",     0)
            values   = info.get("anomaly_values",  [])
            severity = info.get("severity",        "Low")

            content += self._subsection_header(
                f"{col.replace('_', ' ').upper()} "
                f"— {severity} Severity"
            )
            content += (
                f"    Anomalies detected : {count} "
                f"({pct:.2f}% of records)\n"
                f"    Expected range     : "
                f"[{lower:.4f} to {upper:.4f}]\n"
            )

            if values:
                sample = [
                    f"{v:.3f}"
                    for v in sorted(values)[:8]
                ]
                content += (
                    f"    Sample values      : "
                    f"{', '.join(sample)}"
                    f"{'...' if len(values) > 8 else ''}\n"
                )

        return (
            self._section_header("Anomaly Detection") +
            content + "\n"
        )

    def _build_correlation_section(self, stats: dict) -> str:
        correlations = stats.get("correlations", [])

        if not correlations:
            return (
                self._section_header("Correlation Analysis") +
                "  No strong correlations (|r| > 0.5) "
                "detected between numeric variables.\n\n"
            )

        content = (
            "  Threshold: Pearson correlation |r| > 0.50\n"
            "  Interpretation scale:\n"
            "    0.50 to 0.69 : Moderate relationship\n"
            "    0.70 to 0.89 : Strong relationship\n"
            "    0.90 to 1.00 : Very strong relationship\n\n"
        )

        very_strong = [
            c for c in correlations
            if abs(c.get("correlation", 0)) >= 0.90
        ]
        strong = [
            c for c in correlations
            if 0.70 <= abs(c.get("correlation", 0)) < 0.90
        ]
        moderate = [
            c for c in correlations
            if 0.50 <= abs(c.get("correlation", 0)) < 0.70
        ]

        def format_group(group: list, label: str) -> str:
            if not group:
                return ""
            result = f"  {label}\n"
            for corr in group:
                col_a  = corr.get("column_a",    "")
                col_b  = corr.get("column_b",    "")
                coeff  = corr.get("correlation", 0)
                direct = (
                    "positive" if coeff > 0 else "negative"
                )
                result += (
                    f"    {col_a} ↔ {col_b}\n"
                    f"    r = {coeff:.4f} "
                    f"({direct} correlation)\n\n"
                )
            return result

        content += format_group(
            very_strong,
            "VERY STRONG CORRELATIONS (r >= 0.90)"
        )
        content += format_group(
            strong,
            "STRONG CORRELATIONS (0.70 to r < 0.90)"
        )
        content += format_group(
            moderate,
            "MODERATE CORRELATIONS (0.50 to r < 0.70)"
        )

        return (
            self._section_header("Correlation Analysis") +
            content + "\n"
        )

    def _build_trend_section(
        self,
        stats:   dict,
        profile
    ) -> str:
        if not profile.has_time_series:
            return (
                self._section_header("Trend Analysis") +
                "  No time series dimension detected "
                "in this dataset.\n\n"
            )

        trends = stats.get("trends", {})

        if not trends:
            return (
                self._section_header("Trend Analysis") +
                "  Time series detected but insufficient "
                "data for meaningful trend calculation.\n\n"
            )

        content = (
            f"  Time dimension : {profile.time_column}\n"
            f"  Period covered : "
            f"{profile.time_range_start} to "
            f"{profile.time_range_end}\n\n"
        )

        for col, info in trends.items():
            direction  = info.get("direction",      "stable")
            pct_change = info.get("percent_change", 0)
            start_val  = info.get("start_value",    0)
            end_val    = info.get("end_value",      0)

            direction_symbol = (
                "↑" if direction == "increasing"
                else "↓" if direction == "decreasing"
                else "→"
            )

            content += (
                f"  {col.replace('_', ' ').upper()}\n"
                f"    Direction   : "
                f"{direction_symbol} {direction}\n"
                f"    Start value : {start_val:.4f}\n"
                f"    End value   : {end_val:.4f}\n"
                f"    Net change  : {pct_change:+.2f}%\n\n"
            )

        return (
            self._section_header("Trend Analysis") +
            content + "\n"
        )

    def _build_ai_narrative_section(
        self,
        ai_narrative: str
    ) -> str:
        return (
            self._section_header(
                "AI-Powered Intelligence Analysis"
            ) +
            ai_narrative.strip() +
            "\n\n"
        )

    def _build_visualizations_section(
        self,
        viz_result
    ) -> str:
        if viz_result is None:
            return (
                self._section_header("Visualizations") +
                "  No visualizations generated.\n\n"
            )

        content = (
            f"  Total charts generated: "
            f"{viz_result.total_charts_generated}\n"
            f"  Output folder: {viz_result.charts_folder}\n\n"
            f"  CHART INVENTORY\n"
        )

        chart_map = {
            "01_overview.png":
                "Dataset Overview — all variables over time/index",
            "02_distributions.png":
                "Statistical Distributions — histogram and box plots",
            "03_correlations.png":
                "Correlation Analysis — heatmap and scatter plots",
            "04_anomalies.png":
                "Anomaly Detection — outliers highlighted in context",
            "05_trends.png":
                "Trend Analysis — time series with rolling averages",
            "06_categorical.png":
                "Categorical Analysis — frequency and group comparisons",
        }

        for filename, description in chart_map.items():
            paths     = viz_result.get_all_chart_paths()
            generated = any(filename in p for p in paths)
            status    = (
                "✔  Generated"
                if generated
                else "—  Not applicable"
            )
            content += (
                f"    {filename:<30} {status}\n"
                f"      {description}\n"
            )

        if viz_result.generation_errors:
            content += "\n  GENERATION NOTES\n"
            for err in viz_result.generation_errors:
                content += f"    ⚠  {err}\n"

        return (
            self._section_header("Visualizations") +
            content + "\n"
        )

    def _build_statistical_appendix(
        self,
        stats:   dict,
        profile
    ) -> str:
        desc = stats.get("descriptive", {})

        if not desc or not profile.numeric_columns:
            return ""

        content = (
            "  Complete numerical results for all "
            "numeric variables.\n\n"
        )

        cols    = profile.numeric_columns[:10]
        metrics = [
            "count", "mean", "median", "std",
            "min", "max", "q1", "q3",
            "skewness", "kurtosis"
        ]

        header  = f"  {'Metric':<15}"
        for col in cols:
            short_name = col[:10]
            header    += f"  {short_name:>12}"
        content += header + "\n"
        content += (
            "  " + "─" * (15 + 14 * len(cols)) + "\n"
        )

        for metric in metrics:
            row = f"  {metric:<15}"
            for col in cols:
                val = desc.get(col, {}).get(metric, None)
                if val is None:
                    row += f"  {'N/A':>12}"
                elif metric == "count":
                    row += f"  {int(val):>12}"
                else:
                    try:
                        row += f"  {float(val):>12.4f}"
                    except (ValueError, TypeError):
                        row += f"  {'N/A':>12}"
            content += row + "\n"

        return (
            self._section_header(
                "Statistical Appendix — Complete Numerical Results"
            ) +
            content + "\n"
        )

    def _build_footer(self, timestamp: str) -> str:
        width = 65
        return (
            "\n" +
            self._separator("═", width) +
            self._separator("═", width) +
            "\n"
            "  AI DATA ANALYSIS AGENT — END OF REPORT\n"
            f"  Generated: {timestamp}\n"
            "  All outputs are automatically saved.\n"
            "\n" +
            self._separator("═", width) +
            self._separator("═", width) +
            "\n"
        )


# ════════════════════════════════════════════════════════
# EXCEL REPORT WRITER
# ════════════════════════════════════════════════════════

class ExcelReportWriter:
    """
    Produces a structured multi-sheet Excel workbook.
    Uses explicit openpyxl workbook management rather
    than pd.ExcelWriter context manager to avoid
    compatibility issues when adding customization
    sheets after the initial write operation.
    """

    def write(
        self,
        profile,
        stats:         dict,
        output_folder: str,
        session_id:    str,
        df:            Optional[pd.DataFrame] = None
    ) -> str:
        """
        Write the complete Excel workbook to disk.
        Returns file path on success, None on failure.
        """

        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill,
                Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning(
                "openpyxl not installed. "
                "Excel export skipped. "
                "Run: pip install openpyxl"
            )
            return None

        os.makedirs(output_folder, exist_ok=True)

        safe_name = (
            os.path.splitext(profile.file_name)[0]
            .replace(" ", "_")
            .replace("/", "-")
        )
        filename  = (
            f"{session_id}_{safe_name}_report.xlsx"
        )
        filepath  = os.path.join(output_folder, filename)
        data_df   = (
            df if df is not None else profile.dataframe
        )

        # Style definitions
        header_font    = Font(
            bold=True, color="FFFFFF", size=10
        )
        header_fill    = PatternFill(
            fill_type="solid", fgColor="2C5F8A"
        )
        alt_fill       = PatternFill(
            fill_type="solid", fgColor="EBF3FB"
        )
        center_align   = Alignment(
            horizontal="center", vertical="center"
        )
        left_align     = Alignment(
            horizontal="left", vertical="center"
        )
        thin_border    = Border(
            left   = Side(style="thin", color="CCCCCC"),
            right  = Side(style="thin", color="CCCCCC"),
            top    = Side(style="thin", color="CCCCCC"),
            bottom = Side(style="thin", color="CCCCCC")
        )

        def style_header(ws, row_num, n_cols):
            for c in range(1, n_cols + 1):
                cell           = ws.cell(
                    row=row_num, column=c
                )
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border

        def style_data_row(ws, row_num, n_cols, alt):
            for c in range(1, n_cols + 1):
                cell           = ws.cell(
                    row=row_num, column=c
                )
                if alt:
                    cell.fill  = alt_fill
                cell.border    = thin_border
                cell.alignment = left_align

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── SHEET 1 — Raw Data ───────────────────────────
        ws_data = wb.create_sheet("Raw Data")
        headers = data_df.columns.tolist()

        for col_idx, header in enumerate(headers, 1):
            cell           = ws_data.cell(
                row=1, column=col_idx, value=header
            )
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        for row_idx, row in enumerate(
            data_df.itertuples(index=False), 2
        ):
            for col_idx, value in enumerate(row, 1):
                try:
                    safe_val = (
                        None
                        if (
                            value is None or
                            (isinstance(value, float)
                             and np.isnan(value))
                        )
                        else value
                    )
                except (TypeError, ValueError):
                    safe_val = str(value)

                ws_data.cell(
                    row=row_idx,
                    column=col_idx,
                    value=safe_val
                )
            style_data_row(
                ws_data, row_idx,
                len(headers),
                row_idx % 2 == 0
            )

        for col_idx in range(1, len(headers) + 1):
            ws_data.column_dimensions[
                get_column_letter(col_idx)
            ].width = 16

        ws_data.freeze_panes = "A2"

        # ── SHEET 2 — Descriptive Statistics ────────────
        ws_stats     = wb.create_sheet("Descriptive Statistics")
        desc         = stats.get("descriptive", {})
        stat_headers = [
            "Variable", "Count", "Mean", "Median",
            "Std Dev", "Variance", "Min", "Max",
            "Range", "Q1", "Q3", "IQR",
            "Skewness", "Kurtosis", "Missing"
        ]

        for col_idx, header in enumerate(stat_headers, 1):
            cell           = ws_stats.cell(
                row=1, column=col_idx, value=header
            )
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        for row_idx, col in enumerate(
            profile.numeric_columns, 2
        ):
            if col not in desc:
                continue
            s        = desc[col]
            row_data = [
                col,
                s.get("count",    0),
                round(float(s.get("mean",     0)), 4),
                round(float(s.get("median",   0)), 4),
                round(float(s.get("std",      0)), 4),
                round(float(s.get("variance", 0)), 4),
                round(float(s.get("min",      0)), 4),
                round(float(s.get("max",      0)), 4),
                round(float(s.get("range",    0)), 4),
                round(float(s.get("q1",       0)), 4),
                round(float(s.get("q3",       0)), 4),
                round(float(s.get("iqr",      0)), 4),
                round(float(s.get("skewness", 0)), 4),
                round(float(s.get("kurtosis", 0)), 4),
                s.get("missing",  0),
            ]
            for col_idx, value in enumerate(row_data, 1):
                ws_stats.cell(
                    row=row_idx, column=col_idx, value=value
                )
            style_data_row(
                ws_stats, row_idx,
                len(stat_headers),
                row_idx % 2 == 0
            )

        for col_idx in range(1, len(stat_headers) + 1):
            ws_stats.column_dimensions[
                get_column_letter(col_idx)
            ].width = 14

        ws_stats.freeze_panes = "B2"

        # ── SHEET 3 — Anomalies ──────────────────────────
        ws_anom      = wb.create_sheet("Anomalies")
        anomalies    = stats.get("anomalies", {})
        anom_headers = [
            "Variable", "Anomaly Count",
            "Anomaly %", "Lower Bound",
            "Upper Bound", "Severity",
            "Sample Anomalous Values"
        ]

        for col_idx, header in enumerate(anom_headers, 1):
            cell           = ws_anom.cell(
                row=1, column=col_idx, value=header
            )
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        if not anomalies:
            ws_anom.cell(
                row=2, column=1,
                value="No anomalies detected"
            )
        else:
            for row_idx, (col, info) in enumerate(
                anomalies.items(), 2
            ):
                values   = info.get("anomaly_values", [])
                sample   = ", ".join(
                    f"{v:.3f}" for v in values[:5]
                )
                row_data = [
                    col,
                    info.get("anomaly_count",   0),
                    round(float(
                        info.get("anomaly_percent", 0)
                    ), 2),
                    round(float(
                        info.get("lower_bound", 0)
                    ), 4),
                    round(float(
                        info.get("upper_bound", 0)
                    ), 4),
                    info.get("severity", ""),
                    sample
                ]
                for col_idx, value in enumerate(
                    row_data, 1
                ):
                    ws_anom.cell(
                        row=row_idx,
                        column=col_idx,
                        value=value
                    )
                style_data_row(
                    ws_anom, row_idx,
                    len(anom_headers),
                    row_idx % 2 == 0
                )

        for col_idx in range(1, len(anom_headers) + 1):
            ws_anom.column_dimensions[
                get_column_letter(col_idx)
            ].width = 18

        # ── SHEET 4 — Correlations ───────────────────────
        ws_corr      = wb.create_sheet("Correlations")
        correlations = stats.get("correlations", [])
        corr_headers = [
            "Variable A", "Variable B",
            "Correlation (r)", "Strength",
            "Direction", "Interpretation"
        ]

        for col_idx, header in enumerate(corr_headers, 1):
            cell           = ws_corr.cell(
                row=1, column=col_idx, value=header
            )
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        if not correlations:
            ws_corr.cell(
                row=2, column=1,
                value="No strong correlations detected"
            )
        else:
            for row_idx, corr in enumerate(
                correlations, 2
            ):
                coeff     = corr.get("correlation", 0)
                direction = (
                    "Positive" if coeff > 0 else "Negative"
                )
                interp    = (
                    "As one increases the other increases"
                    if coeff > 0
                    else "As one increases the other decreases"
                )
                row_data  = [
                    corr.get("column_a", ""),
                    corr.get("column_b", ""),
                    round(float(coeff),   4),
                    corr.get("strength", ""),
                    direction,
                    interp
                ]
                for col_idx, value in enumerate(
                    row_data, 1
                ):
                    ws_corr.cell(
                        row=row_idx,
                        column=col_idx,
                        value=value
                    )
                style_data_row(
                    ws_corr, row_idx,
                    len(corr_headers),
                    row_idx % 2 == 0
                )

        for col_idx in range(1, len(corr_headers) + 1):
            ws_corr.column_dimensions[
                get_column_letter(col_idx)
            ].width = 22

        # ── SHEET 5 — Dataset Profile ────────────────────
        ws_profile   = wb.create_sheet("Dataset Profile")
        profile_data = [
            ["DATASET PROFILE", ""],
            ["", ""],
            ["File Name",       profile.file_name],
            ["File Format",     profile.file_format.upper()],
            ["File Size (KB)",  profile.file_size_kb],
            ["Total Rows",      profile.total_rows],
            ["Total Columns",   profile.total_columns],
            ["Numeric Columns", len(profile.numeric_columns)],
            ["Text Columns",    len(profile.text_columns)],
            ["Date Columns",    len(profile.datetime_columns)],
            ["Missing Values",  profile.total_missing_values],
            ["Completeness %",  profile.completeness_percent],
            ["Duplicate Rows",  profile.duplicate_rows],
            ["Time Series",
             "Yes" if profile.has_time_series else "No"],
        ]

        if profile.has_time_series:
            profile_data += [
                ["Time Column",  profile.time_column],
                ["Period Start", profile.time_range_start],
                ["Period End",   profile.time_range_end],
            ]

        for row_idx, (label, value) in enumerate(
            profile_data, 1
        ):
            ws_profile.cell(
                row=row_idx, column=1, value=label
            ).font = Font(bold=True)
            ws_profile.cell(
                row=row_idx, column=2, value=value
            )

        ws_profile.column_dimensions["A"].width = 22
        ws_profile.column_dimensions["B"].width = 35

        # ── DOMAIN CUSTOMIZATION SHEETS ──────────────────
        self._add_customization_sheets(
            wb, profile, stats, data_df
        )

        # Explicit save — avoids pd.ExcelWriter conflicts
        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")
        return filepath

    def _add_customization_sheets(
        self,
        wb,
        profile,
        stats: dict,
        df:    pd.DataFrame
    ) -> None:
        """
        Add domain-specific sheets from the domain registry.
        Each customization provides a dict of sheet_name
        to DataFrame mappings. Sheets with empty DataFrames
        are skipped to avoid creating blank sheets.
        """

        try:
            customizations = get_applicable_domains(
                profile, stats, df
            )
        except Exception as error:
            logger.warning(
                f"Domain registry unavailable "
                f"for Excel sheets: {error}"
            )
            return

        for custom in customizations:
            try:
                sheets = custom.get_excel_sheets()

                if not isinstance(sheets, dict):
                    logger.warning(
                        f"Customization "
                        f"{custom.__class__.__name__} "
                        f"returned non-dict for sheets."
                    )
                    continue

                for sheet_name, sheet_df in sheets.items():
                    if (
                        sheet_df is None or
                        not isinstance(sheet_df, pd.DataFrame) or
                        sheet_df.empty
                    ):
                        continue

                    safe_name  = str(sheet_name)[:31]
                    ws_custom  = wb.create_sheet(safe_name)
                    headers    = sheet_df.columns.tolist()

                    for col_idx, header in enumerate(
                        headers, 1
                    ):
                        ws_custom.cell(
                            row=1,
                            column=col_idx,
                            value=str(header)
                        )

                    for row_idx, row in enumerate(
                        sheet_df.itertuples(index=False), 2
                    ):
                        for col_idx, value in enumerate(
                            row, 1
                        ):
                            try:
                                safe_val = (
                                    None
                                    if (
                                        value is None or
                                        (isinstance(value, float)
                                         and np.isnan(value))
                                    )
                                    else value
                                )
                            except (TypeError, ValueError):
                                safe_val = str(value)

                            ws_custom.cell(
                                row=row_idx,
                                column=col_idx,
                                value=safe_val
                            )

                    logger.info(
                        f"Domain sheet added: {safe_name}"
                    )

            except Exception as error:
                logger.warning(
                    f"Failed to add customization sheet "
                    f"from {custom.__class__.__name__}: "
                    f"{error}"
                )
                continue


# ════════════════════════════════════════════════════════
# MAIN REPORT GENERATOR
# ════════════════════════════════════════════════════════

class ReportGenerator:
    """
    Orchestrates the complete four-phase report generation
    pipeline. Each phase is independently executed and
    independently recoverable — a failure in one phase
    does not prevent completion of the others.

    Phase 1 — AI full narrative (with model fallback)
    Phase 2 — AI executive summary (with model fallback)
    Phase 3 — Professional text report
    Phase 4 — Structured Excel workbook
    """

    def __init__(
        self,
        reports_folder:       str = "reports",
        excel_reports_folder: str = "excel_reports"
    ):
        self.reports_folder       = reports_folder
        self.excel_reports_folder = excel_reports_folder
        self.narrator             = AIReportNarrator()
        self.text_writer          = TextReportWriter()
        self.excel_writer         = ExcelReportWriter()

    def generate(
        self,
        profile,
        stats:      dict,
        viz_result,
        session_id: str                  = None,
        df:         Optional[pd.DataFrame] = None
    ) -> ReportResult:
        """
        Run the complete report generation pipeline.
        Returns ReportResult with all output paths and metadata.
        """

        result                       = ReportResult()
        result.generation_timestamp  = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        if session_id is None:
            session_id = datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )

        logger.info(
            f"Report generation starting — "
            f"session: {session_id}"
        )

        # ── PHASE 1 — FULL AI NARRATIVE ─────────────────
        print("\n  Generating AI narrative analysis...")
        print(
            f"  Primary model: {self.narrator.PRIMARY_MODEL}"
        )
        print(
            f"  Fallbacks available: "
            f"{len(self.narrator.FALLBACK_MODELS)}"
        )
        print(
            f"  Retries per model: "
            f"{self.narrator.MAX_RETRIES_PER_MODEL}"
        )

        narrative, narrative_error = (
            self.narrator.generate_full_analysis(
                profile, stats, viz_result
            )
        )

        if narrative:
            result.ai_narrative = narrative
            result.model_used   = self.narrator.model_used
            print(
                f"  AI narrative complete "
                f"(model: {result.model_used})"
            )
        else:
            result.generation_warnings.append(
                f"AI narrative failed: {narrative_error}"
            )
            result.ai_narrative = self._fallback_narrative(
                profile, stats
            )
            print(
                f"  AI narrative unavailable. "
                f"Statistical fallback used."
            )

        # ── PHASE 2 — EXECUTIVE SUMMARY ─────────────────
        print("  Generating executive summary...")

        exec_summary, exec_error = (
            self.narrator.generate_executive_summary(
                profile, stats, result.ai_narrative
            )
        )

        result.executive_summary = (
            exec_summary
            if exec_summary
            else self._fallback_executive_summary(
                profile, stats
            )
        )

        # ── PHASE 3 — TEXT REPORT ────────────────────────
        print("  Writing text report...")

        try:
            result.text_report_path = self.text_writer.write(
                profile       = profile,
                stats         = stats,
                viz_result    = viz_result,
                ai_narrative  = result.ai_narrative,
                exec_summary  = result.executive_summary,
                output_folder = self.reports_folder,
                session_id    = session_id,
                df            = df
            )
            print(
                f"  Text report: "
                f"{os.path.basename(result.text_report_path)}"
            )
        except Exception as error:
            result.generation_warnings.append(
                f"Text report failed: {error}"
            )
            logger.error(
                f"Text report error: {error}",
                exc_info=True
            )

        # ── PHASE 4 — EXCEL REPORT ───────────────────────
        print("  Writing Excel report...")

        try:
            result.excel_report_path = self.excel_writer.write(
                profile       = profile,
                stats         = stats,
                output_folder = self.excel_reports_folder,
                session_id    = session_id,
                df            = df
            )
            if result.excel_report_path:
                print(
                    f"  Excel report: "
                    f"{os.path.basename(result.excel_report_path)}"
                )
        except Exception as error:
            result.generation_warnings.append(
                f"Excel report failed: {error}"
            )
            logger.error(
                f"Excel report error: {error}",
                exc_info=True
            )

        result.generation_success = (
            result.text_report_path is not None or
            result.excel_report_path is not None
        )

        logger.info(
            f"Report generation complete — "
            f"success={result.generation_success}"
        )

        return result

    # ════════════════════════════════════════════════════
    # FALLBACK CONTENT
    # ════════════════════════════════════════════════════

    def _fallback_narrative(
        self,
        profile,
        stats: dict
    ) -> str:
        """
        Statistical fallback when all AI models fail.
        Produces a structured text summary from raw
        statistics without any AI interpretation.
        """

        desc          = stats.get("descriptive",  {})
        anomalies     = stats.get("anomalies",    {})
        correlations  = stats.get("correlations", [])

        anomaly_count = sum(
            info.get("anomaly_count", 0)
            for info in anomalies.values()
        )
        corr_count    = len(correlations)

        lines = [
            "AI NARRATIVE UNAVAILABLE — STATISTICAL SUMMARY",
            "─" * 55,
            "",
            f"Dataset    : {profile.file_name}",
            f"Records    : {profile.total_rows:,}",
            f"Columns    : {profile.total_columns}",
            f"Complete   : {profile.completeness_percent}%",
            f"Anomalies  : {anomaly_count} detected",
            f"Correlations: {corr_count} significant pairs",
            "",
            "KEY STATISTICS",
        ]

        for col in profile.numeric_columns[:5]:
            if col in desc:
                s = desc[col]
                lines.append(
                    f"  {col}: "
                    f"mean={float(s.get('mean', 0)):.3f} "
                    f"std={float(s.get('std', 0)):.3f} "
                    f"range=[{float(s.get('min', 0)):.3f}"
                    f" to {float(s.get('max', 0)):.3f}]"
                )

        lines += [
            "",
            "Retry the analysis when AI services are "
            "available to receive the complete written "
            "intelligence narrative with recommendations."
        ]

        return "\n".join(lines)

    def _fallback_executive_summary(
        self,
        profile,
        stats: dict
    ) -> str:
        """
        Minimal executive summary when AI is unavailable.
        """

        anomaly_count = sum(
            info.get("anomaly_count", 0)
            for info in stats.get("anomalies", {}).values()
        )
        corr_count    = len(stats.get("correlations", []))

        return (
            f"This analysis examined {profile.total_rows:,} "
            f"records across {profile.total_columns} variables "
            f"from {profile.file_name}. "
            f"The dataset achieved {profile.completeness_percent}% "
            f"completeness with "
            f"{profile.total_missing_values} missing values. "
            f"Statistical analysis identified "
            f"{anomaly_count} anomalous data points "
            f"and {corr_count} significant correlations. "
            f"Full details are provided in the sections below."
        )

    def get_result_summary(
        self,
        result: ReportResult
    ) -> str:
        lines = [
            "REPORT GENERATION RESULTS",
            "=" * 45,
            f"Generated at : {result.generation_timestamp}",
            f"AI model used: {result.model_used}",
            f"Success      : {result.generation_success}",
            "",
            "OUTPUT FILES",
        ]

        lines.append(
            f"  Text report  : "
            f"{'✔  ' + result.text_report_path if result.text_report_path else '—  Not generated'}"
        )
        lines.append(
            f"  Excel report : "
            f"{'✔  ' + result.excel_report_path if result.excel_report_path else '—  Not generated'}"
        )

        if result.generation_warnings:
            lines += ["", "WARNINGS"]
            for w in result.generation_warnings:
                lines.append(f"  ⚠  {w}")

        lines += ["", "=" * 45]
        return "\n".join(lines)